from flask import Flask, request, jsonify, send_file, render_template
from flask_cors import CORS
import requests
import pandas as pd
import os
from dotenv import load_dotenv
from datetime import datetime
import io
import re
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Load environment variables
load_dotenv()

app = Flask(__name__)
CORS(app)

class IoCSearcher:
    def __init__(self):
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        
    def search_cve(self, cve_id):
        try:
            # Using NIST NVD API
            base_url = f"https://services.nvd.nist.gov/rest/json/cves/2.0"
            params = {
                'cveId': cve_id
            }
            response = requests.get(base_url, headers=self.headers, params=params)
            
            if response.status_code == 200:
                data = response.json()
                if data['vulnerabilities']:
                    cve_data = data['vulnerabilities'][0]['cve']
                    descriptions = cve_data.get('descriptions', [])
                    description = next((d['value'] for d in descriptions if d['lang'] == 'en'), 'No description available')
                    
                    metrics = cve_data.get('metrics', {}).get('cvssMetricV31', [{}])[0].get('cvssData', {})
                    
                    return {
                        'type': 'CVE',
                        'value': cve_id,
                        'source': 'NIST NVD',
                        'description': description,
                        'base_score': metrics.get('baseScore', 'N/A'),
                        'severity': metrics.get('baseSeverity', 'N/A'),
                        'published': cve_data.get('published', 'N/A'),
                        'last_modified': cve_data.get('lastModified', 'N/A')
                    }
            return None
        except Exception as e:
            print(f"Error searching CVE: {str(e)}")
            return None
        
    def search_abuseipdb(self, keyword):
        api_key = os.getenv('ABUSEIPDB_API_KEY')
        if not api_key:
            return []
            
        base_url = "https://api.abuseipdb.com/api/v2/check"
        results = []
        
        try:
            params = {
                'ipAddress': keyword,
                'maxAgeInDays': '90'
            }
            headers = {
                'Key': api_key,
                'Accept': 'application/json'
            }
            
            response = requests.get(base_url, params=params, headers=headers)
            if response.status_code == 200:
                data = response.json()
                if data['data']:
                    results.append({
                        'type': 'IP',
                        'value': keyword,
                        'source': 'AbuseIPDB',
                        'confidence': data['data']['abuseConfidenceScore'],
                        'country': data['data']['countryCode'],
                        'last_reported': data['data']['lastReportedAt']
                    })
        except Exception as e:
            print(f"Error searching AbuseIPDB: {str(e)}")
            
        return results
        
    def search_virustotal(self, keyword):
        api_key = os.getenv('VIRUSTOTAL_API_KEY')
        if not api_key:
            return []
            
        base_url = "https://www.virustotal.com/api/v3"
        results = []
        
        try:
            if '.' in keyword:  # Could be IP or domain
                endpoint = f"{base_url}/ip_addresses/{keyword}"
                response = requests.get(endpoint, headers={'x-apikey': api_key})
                if response.status_code == 200:
                    data = response.json()
                    if 'data' in data:
                        results.append({
                            'type': 'IP/Domain',
                            'value': keyword,
                            'source': 'VirusTotal',
                            'malicious': data['data']['attributes'].get('last_analysis_stats', {}).get('malicious', 0),
                            'last_analysis': data['data']['attributes'].get('last_analysis_date', '')
                        })
            else:  # Assume it's a hash
                endpoint = f"{base_url}/files/{keyword}"
                response = requests.get(endpoint, headers={'x-apikey': api_key})
                if response.status_code == 200:
                    data = response.json()
                    if 'data' in data:
                        results.append({
                            'type': 'Hash',
                            'value': keyword,
                            'source': 'VirusTotal',
                            'malicious': data['data']['attributes'].get('last_analysis_stats', {}).get('malicious', 0),
                            'sha256': data['data']['attributes'].get('sha256', ''),
                            'sha1': data['data']['attributes'].get('sha1', ''),
                            'md5': data['data']['attributes'].get('md5', ''),
                            'last_analysis': data['data']['attributes'].get('last_analysis_date', '')
                        })
        except Exception as e:
            print(f"Error searching VirusTotal: {str(e)}")
            
        return results
        
    def search_threatfox(self, keyword):
        api_key = os.getenv('THREATFOX_API_KEY')
        if not api_key:
            return []
            
        base_url = "https://threatfox.abuse.ch/api/v1/"
        results = []
        
        try:
            payload = {
                'query': 'search_ioc',
                'search_term': keyword
            }
            headers = {
                'API-KEY': api_key,
                'Content-Type': 'application/json'
            }
            
            response = requests.post(base_url, json=payload, headers=headers)
            if response.status_code == 200:
                data = response.json()
                if data.get('data'):
                    for ioc in data['data']:
                        results.append({
                            'type': ioc.get('ioc_type', 'Unknown'),
                            'value': ioc.get('ioc', keyword),
                            'source': 'ThreatFox',
                            'malware': ioc.get('malware', 'Unknown'),
                            'first_seen': ioc.get('first_seen', '')
                        })
        except Exception as e:
            print(f"Error searching ThreatFox: {str(e)}")
            
        return results

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/search', methods=['POST'])
def search_iocs():
    data = request.json
    feeds = data.get('feeds', [])
    results = {}
    
    searcher = IoCSearcher()
    
    for feed in feeds:
        feed_title = feed.get('title', 'Untitled Feed')
        keywords = [k.strip() for k in feed.get('iocs', '').split(',')]
        feed_results = []
        
        for keyword in keywords:
            if keyword.startswith('CVE-'):
                cve_result = searcher.search_cve(keyword)
                if cve_result:
                    feed_results.append(cve_result)
            else:
                feed_results.extend(searcher.search_abuseipdb(keyword))
                feed_results.extend(searcher.search_virustotal(keyword))
                feed_results.extend(searcher.search_threatfox(keyword))
        
        results[feed_title] = feed_results
    
    return jsonify(results)

@app.route('/api/export', methods=['POST'])
def export_to_excel():
    data = request.json
    feeds_results = data.get('results', {})
    
    if not feeds_results:
        return jsonify({'error': 'No data to export'}), 400
        
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for feed_title, results in feeds_results.items():
            if results:
                df = pd.DataFrame(results)
                # Clean sheet name (Excel has restrictions on sheet names)
                sheet_name = re.sub(r'[\[\]\*\?:/\\]', '', feed_title)[:31]
                
                # Write the DataFrame starting from row 2 (leaving space for title)
                df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
                
                # Get the worksheet to apply formatting
                worksheet = writer.sheets[sheet_name]
                
                # Add title in the first row
                last_column_letter = get_column_letter(len(df.columns))
                worksheet.merge_cells(f'A1:{last_column_letter}1')  # Merge cells for title
                title_cell = worksheet['A1']
                title_cell.value = feed_title
                title_cell.font = Font(size=18)
                title_cell.alignment = Alignment(horizontal='center')
                
                # Adjust column widths
                for idx, column in enumerate(df.columns, start=1):
                    column_letter = get_column_letter(idx)
                    # Get max length of column content
                    max_length = max(
                        df[column].astype(str).apply(len).max(),
                        len(str(column))
                    )
                    adjusted_width = min(max_length + 2, 50)  # Cap width at 50
                    worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'threat_bulletin_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True) 
